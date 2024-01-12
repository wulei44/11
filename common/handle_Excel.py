import openpyxl


class Login_Excel():
    def __init__(self, filename, sheetname):
        """

        :param filename: 用例文件地址
        :param sheetname: 用例表单
        """
        self.filename = filename
        self.sheetname = sheetname

    def login_excel_read(self):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        title = [i.value for i in res[0]]

        case = []
        for itme in res[1:]:
            data = [i.value for i in itme]
            dic = dict(zip(title, data))
            case.append(dic)
        return case

        # 写入数据的方法

    def write_data(self, row, column, value):
        """

        :param row: 写入excel表的行号
        :param column: 写入excel表的列
        :param value: 写入excel表的值
        :return:
        """
        # 读取Excel表格
        workbook = openpyxl.load_workbook(self.filename)
        # 选中表单
        sh = workbook[self.sheetname]
        # 写入数据
        sh.cell(row=row, column=column, value=value)
        # 重新刷新表格，并保存数据
        workbook.save(self.filename)
